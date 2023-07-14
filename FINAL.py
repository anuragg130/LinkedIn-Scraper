import openpyxl

def store_list_in_excel(data):
    wb = openpyxl.Workbook()
    sheet = wb.active

    row = 1
    for item in data:
        sheet.cell(row, 1).value = item[0]
        for value in item[1:]:
            sheet.cell(row, value + 1).value = value
        row += 1

    wb.save('list_in_excel.xlsx')

if __name__ == '__main__':
    data = [['A', 'B', 'C'], [1, 2, 3], [4, 5, 6], [7, 8, 9]]

    store_list_in_excel(data)